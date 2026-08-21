#!/usr/bin/env python3
"""Fetch DingTalk knowledge-base document metadata or blocks.

This script is read-only. It keeps raw API JSON available because DingTalk block
schemas can vary by document type and tenant capability.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
from pathlib import Path
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.parse import quote, urlencode
from urllib.request import Request, urlopen


API_BASE = "https://api.dingtalk.com"
OAPI_BASE = "https://oapi.dingtalk.com"


class DingTalkError(RuntimeError):
    pass


def load_dotenv(repo_root: Path) -> None:
    env_path = repo_root / ".env"
    if not env_path.exists():
        return
    for raw_line in env_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        os.environ.setdefault(key, value)


def request_json(
    method: str,
    path: str,
    token: str | None = None,
    params: dict[str, Any] | None = None,
    body: dict[str, Any] | None = None,
) -> dict[str, Any]:
    url = f"{API_BASE}{path}"
    if params:
        url = f"{url}?{urlencode({k: v for k, v in params.items() if v is not None})}"
    data = None
    headers = {"Content-Type": "application/json"}
    if token:
        headers["x-acs-dingtalk-access-token"] = token
    if body is not None:
        data = json.dumps(body, ensure_ascii=False).encode("utf-8")

    req = Request(url, data=data, headers=headers, method=method.upper())
    try:
        with urlopen(req, timeout=30) as resp:
            payload = resp.read().decode("utf-8")
    except HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")
        raise DingTalkError(f"HTTP {exc.code} {method} {url}: {detail}") from exc
    except URLError as exc:
        raise DingTalkError(f"Network error calling {method} {url}: {exc}") from exc

    try:
        result = json.loads(payload)
    except json.JSONDecodeError as exc:
        raise DingTalkError(f"Non-JSON response from {method} {url}: {payload[:500]}") from exc

    code = result.get("code") or result.get("errcode")
    if code not in (None, "", 0, "0"):
        msg = result.get("message") or result.get("errmsg") or result
        raise DingTalkError(f"DingTalk API error {code} from {method} {url}: {msg}")
    return result


def request_oapi_json(
    method: str,
    path: str,
    token: str,
    params: dict[str, Any] | None = None,
    body: dict[str, Any] | None = None,
) -> dict[str, Any]:
    query = {"access_token": token}
    if params:
        query.update({k: v for k, v in params.items() if v is not None})
    url = f"{OAPI_BASE}{path}?{urlencode(query)}"
    data = None
    headers = {"Content-Type": "application/json"}
    if body is not None:
        data = json.dumps(body, ensure_ascii=False).encode("utf-8")

    req = Request(url, data=data, headers=headers, method=method.upper())
    try:
        with urlopen(req, timeout=30) as resp:
            payload = resp.read().decode("utf-8")
    except HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")
        raise DingTalkError(f"HTTP {exc.code} {method} {OAPI_BASE}{path}: {detail}") from exc
    except URLError as exc:
        raise DingTalkError(f"Network error calling {method} {OAPI_BASE}{path}: {exc}") from exc

    try:
        result = json.loads(payload)
    except json.JSONDecodeError as exc:
        raise DingTalkError(f"Non-JSON response from {method} {OAPI_BASE}{path}: {payload[:500]}") from exc

    errcode = result.get("errcode")
    if errcode not in (None, 0, "0"):
        errmsg = result.get("errmsg") or result
        raise DingTalkError(f"DingTalk OAPI error {errcode} from {method} {OAPI_BASE}{path}: {errmsg}")
    return result


def unwrap_result(payload: dict[str, Any]) -> Any:
    for key in ("result", "data"):
        if key in payload:
            return payload[key]
    return payload


def get_access_token(app_key: str, app_secret: str) -> str:
    payload = request_json(
        "POST",
        "/v1.0/oauth2/accessToken",
        body={"appKey": app_key, "appSecret": app_secret},
    )
    token = payload.get("accessToken") or unwrap_result(payload).get("accessToken")
    if not token:
        raise DingTalkError(f"accessToken missing in response: {json.dumps(payload, ensure_ascii=False)[:500]}")
    return token


def get_userid_by_mobile(token: str, mobile: str) -> dict[str, Any]:
    return request_oapi_json("GET", "/user/get_by_mobile", token, params={"mobile": mobile})


def get_unionid_by_userid(token: str, userid: str) -> dict[str, Any]:
    return request_oapi_json(
        "POST",
        "/topapi/v2/user/get",
        token,
        body={"userid": userid, "language": "zh_CN"},
    )


def resolve_access_token(args: argparse.Namespace) -> str:
    if args.access_token:
        return args.access_token

    app_key = args.app_key or os.getenv("DINGTALK_APP_KEY")
    app_secret = args.app_secret or os.getenv("DINGTALK_APP_SECRET")
    if app_key and app_secret:
        return get_access_token(app_key, app_secret)

    existing_token = os.getenv("DINGTALK_ACCESS_TOKEN") or os.getenv("DINGTALK_API_TOKEN")
    if existing_token:
        return existing_token
    raise DingTalkError("Missing access token, or app key/app secret for token generation.")


def query_node_by_url(token: str, operator_id: str, url: str) -> dict[str, Any]:
    payload = request_json(
        "POST",
        "/v2.0/wiki/nodes/queryByUrl",
        token=token,
        params={"operatorId": operator_id},
        body={
            "url": url,
            "option": {
                "withStatisticalInfo": True,
                "withPermissionRole": True,
            },
        },
    )
    result = unwrap_result(payload)
    if isinstance(result, dict):
        return result
    raise DingTalkError(f"Unexpected queryByUrl response: {json.dumps(payload, ensure_ascii=False)[:500]}")


def find_first_value(value: Any, keys: tuple[str, ...]) -> Any:
    if isinstance(value, dict):
        for key in keys:
            item = value.get(key)
            if item not in (None, ""):
                return item
        for item in value.values():
            found = find_first_value(item, keys)
            if found not in (None, ""):
                return found
    elif isinstance(value, list):
        for item in value:
            found = find_first_value(item, keys)
            if found not in (None, ""):
                return found
    return None


def resolve_space_url(
    token: str,
    operator_id: str,
    url: str,
) -> tuple[str | None, str | None, dict[str, Any] | None, list[str]]:
    """Resolve a space URL by API first and regex as a last fallback.

    Returns (workspace_id, root_node_id, raw_node).
    """
    raw_node: dict[str, Any] | None = None
    diagnostics: list[str] = []
    workspace_id = extract_workspace_id_from_url(url)
    root_node_id: str | None = None
    try:
        raw_node = query_node_by_url(token, operator_id, url)
        api_workspace_id = find_first_value(raw_node, ("workspaceId", "spaceId"))
        api_root_node_id = find_first_value(raw_node, ("rootNodeId", "rootNodeID"))
        api_node_id = find_first_value(raw_node, ("nodeId", "dentryUuid"))
        if isinstance(api_workspace_id, str) and api_workspace_id:
            workspace_id = api_workspace_id
        if isinstance(api_root_node_id, str) and api_root_node_id:
            root_node_id = api_root_node_id
        elif isinstance(api_node_id, str) and api_node_id:
            root_node_id = api_node_id
        else:
            diagnostics.append(f"queryByUrl returned no root/node id: {json.dumps(raw_node, ensure_ascii=False)[:500]}")
    except DingTalkError as exc:
        diagnostics.append(f"queryByUrl failed: {exc}")

    try:
        url_workspace_slug = extract_workspace_id_from_url(url)
        for workspace in list_workspaces(token, operator_id):
            workspace_url = workspace.get("url")
            workspace_slug = extract_workspace_id_from_url(workspace_url) if isinstance(workspace_url, str) else None
            if workspace_url == url or (url_workspace_slug and workspace_slug == url_workspace_slug):
                api_workspace_id = workspace.get("workspaceId")
                api_root_node_id = workspace.get("rootNodeId")
                if isinstance(api_workspace_id, str) and api_workspace_id:
                    workspace_id = api_workspace_id
                if isinstance(api_root_node_id, str) and api_root_node_id:
                    root_node_id = api_root_node_id
                raw_node = workspace
                break
        else:
            diagnostics.append("workspace list did not contain a URL matching the space URL")
    except DingTalkError as exc:
        diagnostics.append(f"workspace URL match failed: {exc}")
    return workspace_id, root_node_id, raw_node, diagnostics


def get_mine_workspaces(token: str, operator_id: str) -> Any:
    payload = request_json(
        "GET",
        "/v2.0/wiki/mineWorkspaces",
        token=token,
        params={"operatorId": operator_id},
    )
    return unwrap_result(payload)


def get_workspace(token: str, operator_id: str, workspace_id: str) -> dict[str, Any]:
    payload = request_json(
        "GET",
        f"/v2.0/wiki/workspaces/{quote(workspace_id, safe='')}",
        token=token,
        params={"operatorId": operator_id},
    )
    result = unwrap_result(payload)
    if isinstance(result, dict):
        workspace = result.get("workspace") if isinstance(result.get("workspace"), dict) else result
        if isinstance(workspace, dict):
            return workspace
    raise DingTalkError(f"Unexpected workspace response: {json.dumps(payload, ensure_ascii=False)[:500]}")


def list_workspaces(token: str, operator_id: str, max_results: int = 50) -> list[dict[str, Any]]:
    workspaces: list[dict[str, Any]] = []
    next_token: str | None = None
    while True:
        payload = request_json(
            "GET",
            "/v2.0/wiki/workspaces",
            token=token,
            params={
                "operatorId": operator_id,
                "maxResults": max_results,
                "nextToken": next_token,
            },
        )
        result = unwrap_result(payload)
        batch = pick_first_list(result, ("workspaces", "items", "list", "data")) or []
        workspaces.extend(item for item in batch if isinstance(item, dict))
        if not isinstance(result, dict):
            break
        next_token = result.get("nextToken") or result.get("nextCursor")
        has_more = result.get("hasMore")
        if not next_token or has_more is False:
            break
    return workspaces


def resolve_root_node_id(
    token: str,
    operator_id: str,
    workspace_id: str,
    max_results: int = 50,
) -> tuple[str, dict[str, Any] | None]:
    errors: list[str] = []
    try:
        workspace = get_workspace(token, operator_id, workspace_id)
        root_node_id = workspace.get("rootNodeId")
        if isinstance(root_node_id, str) and root_node_id:
            return root_node_id, workspace
        errors.append(f"workspace detail had no rootNodeId: {json.dumps(workspace, ensure_ascii=False)[:500]}")
    except DingTalkError as exc:
        errors.append(f"workspace detail failed: {exc}")

    try:
        for workspace in list_workspaces(token, operator_id, max_results):
            if workspace.get("workspaceId") == workspace_id:
                root_node_id = workspace.get("rootNodeId")
                if isinstance(root_node_id, str) and root_node_id:
                    return root_node_id, workspace
                errors.append(f"workspace list matched but had no rootNodeId: {json.dumps(workspace, ensure_ascii=False)[:500]}")
                break
        else:
            errors.append("workspace list did not contain the target workspaceId")
    except DingTalkError as exc:
        errors.append(f"workspace list failed: {exc}")

    raise DingTalkError(f"rootNodeId not found for workspaceId={workspace_id}. Diagnostics: {' | '.join(errors)}")


def list_nodes(token: str, operator_id: str, parent_node_id: str, max_results: int = 50) -> list[dict[str, Any]]:
    nodes: list[dict[str, Any]] = []
    next_token: str | None = None
    while True:
        payload = request_json(
            "GET",
            "/v2.0/wiki/nodes",
            token=token,
            params={
                "operatorId": operator_id,
                "parentNodeId": parent_node_id,
                "maxResults": max_results,
                "nextToken": next_token,
            },
        )
        result = unwrap_result(payload)
        batch = pick_first_list(result, ("nodes", "items", "list", "data")) or []
        nodes.extend(item for item in batch if isinstance(item, dict))
        if not isinstance(result, dict):
            break
        next_token = result.get("nextToken") or result.get("nextCursor")
        has_more = result.get("hasMore")
        if not next_token or has_more is False:
            break
    return nodes


def list_workspace_nodes(token: str, operator_id: str, workspace_id: str, max_results: int = 50) -> list[dict[str, Any]]:
    root_node_id, _workspace = resolve_root_node_id(token, operator_id, workspace_id, max_results)
    return list_nodes(token, operator_id, root_node_id, max_results)


def read_blocks(token: str, operator_id: str, doc_key: str, max_results: int = 100) -> dict[str, Any]:
    blocks: list[dict[str, Any]] = []
    raw_pages: list[dict[str, Any]] = []
    next_token: str | None = None
    while True:
        payload = request_json(
            "GET",
            f"/v1.0/doc/suites/documents/{quote(doc_key, safe='')}/blocks",
            token=token,
            params={
                "operatorId": operator_id,
                "maxResults": max_results,
                "nextToken": next_token,
            },
        )
        raw_pages.append(payload)
        result = unwrap_result(payload)
        batch = pick_first_list(result, ("blocks", "items", "list", "data")) or []
        blocks.extend(item for item in batch if isinstance(item, dict))
        if not isinstance(result, dict):
            break
        next_token = result.get("nextToken") or result.get("nextCursor")
        has_more = result.get("hasMore")
        if not next_token or has_more is False:
            break
    return {"blocks": blocks, "rawPages": raw_pages}


def list_workbook_sheets(token: str, operator_id: str, workbook_id: str) -> list[dict[str, Any]]:
    payload = request_json(
        "GET",
        f"/v1.0/doc/workbooks/{quote(workbook_id, safe='')}/sheets",
        token=token,
        params={"operatorId": operator_id},
    )
    result = unwrap_result(payload)
    sheets = pick_first_list(result, ("value", "sheets", "items", "list", "data")) or []
    return [sheet for sheet in sheets if isinstance(sheet, dict)]


def get_workbook_sheet(token: str, operator_id: str, workbook_id: str, sheet_id: str) -> dict[str, Any]:
    payload = request_json(
        "GET",
        f"/v1.0/doc/workbooks/{quote(workbook_id, safe='')}/sheets/{quote(sheet_id, safe='')}",
        token=token,
        params={"operatorId": operator_id},
    )
    result = unwrap_result(payload)
    if isinstance(result, dict):
        return result
    raise DingTalkError(f"Unexpected sheet response: {json.dumps(payload, ensure_ascii=False)[:500]}")


def read_workbook_range(
    token: str,
    operator_id: str,
    workbook_id: str,
    sheet_id: str,
    range_address: str,
    select: str = "values,formulas,displayValues",
) -> dict[str, Any]:
    payload = request_json(
        "GET",
        f"/v1.0/doc/workbooks/{quote(workbook_id, safe='')}/sheets/{quote(sheet_id, safe='')}/ranges/{quote(range_address, safe='')}",
        token=token,
        params={"operatorId": operator_id, "select": select},
    )
    result = unwrap_result(payload)
    if isinstance(result, dict):
        return result
    raise DingTalkError(f"Unexpected range response: {json.dumps(payload, ensure_ascii=False)[:500]}")


def is_folder_node(node: dict[str, Any]) -> bool:
    node_type = str(node.get("type") or node.get("nodeType") or node.get("contentType") or "").lower()
    category = str(node.get("category") or node.get("extension") or "").lower()
    return any(term in node_type for term in ("folder", "directory")) or category == "folder"


def is_document_node(node: dict[str, Any]) -> bool:
    node_type = str(node.get("type") or node.get("nodeType") or node.get("contentType") or "").lower()
    category = str(node.get("category") or "").lower()
    extension = str(node.get("extension") or "").lower().lstrip(".")
    if is_folder_node(node):
        return False
    if extension == "adoc":
        return True
    if any(term in node_type for term in ("doc", "alidoc")) and extension in {"", "adoc"}:
        return True
    return category in {"adoc", "doc", "document"}


def is_spreadsheet_node(node: dict[str, Any]) -> bool:
    node_type = str(node.get("type") or node.get("nodeType") or node.get("contentType") or "").lower()
    category = str(node.get("category") or "").lower()
    extension = str(node.get("extension") or "").lower().lstrip(".")
    if is_folder_node(node):
        return False
    return extension == "axls" or category == "axls" or ("spreadsheet" in node_type and extension in {"", "axls"})


def is_aitable_node(node: dict[str, Any]) -> bool:
    extension = str(node.get("extension") or "").lower().lstrip(".")
    category = str(node.get("category") or "").lower()
    return not is_folder_node(node) and (extension == "able" or category == "able")


def node_title(node: dict[str, Any]) -> str:
    for key in ("name", "title", "nodeName", "fileName"):
        value = node.get(key)
        if isinstance(value, str) and value.strip():
            return value.strip()
    return str(node.get("nodeId") or node.get("docKey") or "untitled")


def crawl_nodes(
    token: str,
    operator_id: str,
    root_nodes: list[dict[str, Any]],
    max_results: int,
    depth_limit: int,
) -> list[dict[str, Any]]:
    collected: list[dict[str, Any]] = []
    queue: list[tuple[dict[str, Any], int, str]] = [(node, 0, node_title(node)) for node in root_nodes]
    seen: set[str] = set()

    while queue:
        node, depth, path = queue.pop(0)
        node_id = extract_doc_key(node)
        if node_id and node_id in seen:
            continue
        if node_id:
            seen.add(node_id)
        annotated = dict(node)
        annotated["_path"] = path
        collected.append(annotated)

        if depth >= depth_limit or not is_folder_node(node) or not node_id:
            continue
        children = list_nodes(token, operator_id, node_id, max_results)
        for child in children:
            queue.append((child, depth + 1, f"{path}/{node_title(child)}"))
    return collected


def fetch_documents(
    token: str,
    operator_id: str,
    nodes: list[dict[str, Any]],
    max_results: int,
    limit_docs: int | None,
) -> list[dict[str, Any]]:
    documents: list[dict[str, Any]] = []
    for node in nodes:
        if limit_docs is not None and len(documents) >= limit_docs:
            break
        if not is_document_node(node):
            continue
        doc_key = extract_doc_key(node)
        if not doc_key:
            continue
        try:
            raw = read_blocks(token, operator_id, doc_key, max_results)
            documents.append({"node": node, **raw, "error": None})
        except DingTalkError as exc:
            documents.append({"node": node, "blocks": [], "rawPages": [], "error": str(exc)})
    return documents


def fetch_spreadsheets(
    token: str,
    operator_id: str,
    nodes: list[dict[str, Any]],
    limit_sheets: int | None,
) -> list[dict[str, Any]]:
    spreadsheets: list[dict[str, Any]] = []
    for node in nodes:
        if limit_sheets is not None and len(spreadsheets) >= limit_sheets:
            break
        if not is_spreadsheet_node(node):
            continue
        workbook_id = extract_doc_key(node)
        if not workbook_id:
            continue
        try:
            sheets = list_workbook_sheets(token, operator_id, workbook_id)
            exported_sheets: list[dict[str, Any]] = []
            for sheet in sheets:
                sheet_id = str(sheet.get("id") or sheet.get("sheetId") or "").strip()
                if not sheet_id:
                    continue
                detail = get_workbook_sheet(token, operator_id, workbook_id, sheet_id)
                range_address = workbook_non_empty_range(detail)
                range_data = {"values": [], "formulas": [], "displayValues": []}
                if range_address:
                    range_data = read_workbook_range(token, operator_id, workbook_id, sheet_id, range_address)
                exported_sheets.append(
                    {
                        "sheet": sheet,
                        "detail": detail,
                        "rangeAddress": range_address,
                        "range": range_data,
                    }
                )
            spreadsheets.append({"node": node, "workbookId": workbook_id, "sheets": exported_sheets, "error": None})
        except DingTalkError as exc:
            spreadsheets.append({"node": node, "workbookId": workbook_id, "sheets": [], "error": str(exc)})
    return spreadsheets


def workbook_non_empty_range(sheet_detail: dict[str, Any]) -> str | None:
    last_row = parse_non_negative_int(sheet_detail.get("lastNonEmptyRow"))
    last_column = parse_non_negative_int(sheet_detail.get("lastNonEmptyColumn"))
    if last_row is None or last_column is None:
        return None
    return f"A1:{column_index_to_letters(last_column + 1)}{last_row + 1}"


def parse_non_negative_int(value: Any) -> int | None:
    try:
        number = int(value)
    except (TypeError, ValueError):
        return None
    return number if number >= 0 else None


def column_index_to_letters(index: int) -> str:
    if index <= 0:
        raise ValueError("Column index must be 1-based and positive.")
    letters: list[str] = []
    while index:
        index, remainder = divmod(index - 1, 26)
        letters.append(chr(65 + remainder))
    return "".join(reversed(letters))


def pick_first_list(value: Any, keys: tuple[str, ...]) -> list[Any] | None:
    if isinstance(value, list):
        return value
    if not isinstance(value, dict):
        return None
    for key in keys:
        item = value.get(key)
        if isinstance(item, list):
            return item
    for item in value.values():
        found = pick_first_list(item, keys)
        if found is not None:
            return found
    return None


def extract_doc_key(node: dict[str, Any], fallback_url: str | None = None) -> str | None:
    for key in ("docKey", "nodeId", "documentId", "dentryUuid", "resourceId"):
        value = node.get(key)
        if isinstance(value, str) and value:
            return value
    for key in ("url", "docUrl", "resourceUrl"):
        value = node.get(key)
        if isinstance(value, str):
            matched = extract_node_id_from_url(value)
            if matched:
                return matched
    if fallback_url:
        return extract_node_id_from_url(fallback_url)
    return None


def extract_node_id_from_url(url: str) -> str | None:
    match = re.search(r"/i/nodes/([^/?#]+)", url)
    if match:
        return match.group(1)
    match = re.search(r"[?&]dentryKey=([^&#]+)", url)
    if match:
        return match.group(1)
    return None


def extract_workspace_id_from_url(url: str) -> str | None:
    match = re.search(r"/i/spaces/([^/?#]+)/overview", url)
    if match:
        return match.group(1)
    return None


def collect_text(value: Any) -> str:
    if isinstance(value, str):
        return value
    if isinstance(value, list):
        return "".join(collect_text(item) for item in value)
    if not isinstance(value, dict):
        return ""

    text_parts: list[str] = []
    for key in ("text", "plainText", "content", "value"):
        item = value.get(key)
        if isinstance(item, str):
            text_parts.append(item)
    for key in ("paragraph", "heading", "elements", "children", "textRuns", "runs", "cells", "rows"):
        item = value.get(key)
        if isinstance(item, (dict, list)):
            text_parts.append(collect_text(item))
    return "".join(text_parts)


def render_markdown_table(cells: list[Any]) -> str:
    rows = [[collect_text(cell).replace("\n", "<br>").strip() for cell in row] for row in cells if isinstance(row, list)]
    if not rows:
        return ""
    col_count = max(len(row) for row in rows)
    normalized = [row + [""] * (col_count - len(row)) for row in rows]
    header = normalized[0]
    separator = ["---"] * col_count
    body = normalized[1:]

    def render_row(row: list[str]) -> str:
        escaped = [cell.replace("|", "\\|") for cell in row]
        return "| " + " | ".join(escaped) + " |"

    return "\n".join([render_row(header), render_row(separator), *[render_row(row) for row in body]])


def render_attachment(block: dict[str, Any]) -> str:
    attachment = block.get("attachment")
    if not isinstance(attachment, dict):
        return ""
    resource_id = str(attachment.get("resourceId") or "").strip()
    name = str(attachment.get("name") or resource_id or "attachment").strip()
    mime_type = str(attachment.get("type") or "").lower()
    size = attachment.get("size")
    suffix = f", {size} bytes" if isinstance(size, int) else ""
    target = f"dingtalk-resource://{resource_id}" if resource_id else ""
    safe_name = name.replace("[", "\\[").replace("]", "\\]")
    if mime_type.startswith("image/"):
        return f"![{safe_name}]({target})"
    return f"[附件: {safe_name}{suffix}]({target})"


def block_to_markdown(block: dict[str, Any]) -> str:
    block_type = str(block.get("type") or block.get("blockType") or block.get("elementType") or "").lower()
    if block_type == "attachment":
        return render_attachment(block)

    if block_type == "unknown":
        block_id = block.get("id") or block.get("blockId") or ""
        return f"<!-- unsupported DingTalk block: unknown id={block_id} -->"

    if block_type == "callout":
        text = collect_text(block).strip()
        return f"> {text}" if text else "<!-- unsupported DingTalk block: callout -->"

    if block_type == "blockquote":
        text = collect_text(block).strip()
        return "\n".join(f"> {line}" for line in text.splitlines()) if text else ""

    if block_type == "table" and isinstance(block.get("table"), dict):
        cells = block["table"].get("cells")
        if isinstance(cells, list):
            return render_markdown_table(cells)

    if block_type == "heading" and isinstance(block.get("heading"), dict):
        heading = block["heading"]
        text = collect_text(heading).strip()
        if not text:
            return ""
        level_text = str(heading.get("level") or "")
        level_match = re.search(r"([1-6])", level_text)
        level = int(level_match.group(1)) if level_match else 2
        return f"{'#' * level} {text}"

    text = collect_text(block).strip()
    if not text:
        return ""
    if "heading" in block_type or block_type in {"h1", "h2", "h3"}:
        level = 2
        level_match = re.search(r"([1-6])", block_type)
        if level_match:
            level = int(level_match.group(1))
        return f"{'#' * level} {text}"
    if "bullet" in block_type or "list" in block_type:
        return f"- {text}"
    if "code" in block_type:
        return f"```\n{text}\n```"
    return text


def blocks_to_markdown(blocks: list[dict[str, Any]]) -> str:
    def order_key(item: dict[str, Any]) -> tuple[int, str]:
        order = item.get("index", item.get("sequence", item.get("sort", 0)))
        try:
            return (int(order), str(item.get("id") or item.get("blockId") or ""))
        except (TypeError, ValueError):
            return (0, str(item.get("id") or item.get("blockId") or ""))

    lines = [block_to_markdown(block) for block in sorted(blocks, key=order_key)]
    return "\n\n".join(line for line in lines if line).strip() + "\n"


def write_output(path: str | None, content: str) -> None:
    if path:
        output_path = Path(path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(content, encoding="utf-8")
    else:
        print(content)


def write_crawl_outputs(
    output_dir: str,
    documents: list[dict[str, Any]],
    spreadsheets: list[dict[str, Any]],
    include_raw: bool,
) -> None:
    base = Path(output_dir)
    base.mkdir(parents=True, exist_ok=True)
    index: dict[str, list[dict[str, Any]]] = {"documents": [], "spreadsheets": []}
    for idx, document in enumerate(documents, start=1):
        node = document.get("node", {})
        title = sanitize_filename(node_title(node))
        stem = f"{idx:03d}_{title}"
        markdown_path = base / f"{stem}.md"
        raw_path = base / f"{stem}.blocks.json"
        if document.get("error"):
            markdown_path.write_text(f"# {node_title(node)}\n\nERROR: {document['error']}\n", encoding="utf-8")
        else:
            markdown_path.write_text(blocks_to_markdown(document.get("blocks", [])), encoding="utf-8")
        if include_raw:
            raw_path.write_text(json.dumps(document, ensure_ascii=False, indent=2), encoding="utf-8")
        index["documents"].append(
            {
                "title": node_title(node),
                "path": node.get("_path"),
                "nodeId": node.get("nodeId"),
                "docKey": node.get("docKey"),
                "blocks": len(document.get("blocks", [])),
                "error": document.get("error"),
                "markdown": str(markdown_path),
                "raw": str(raw_path) if include_raw else None,
            }
        )

    spreadsheet_base = base / "spreadsheets"
    for idx, spreadsheet in enumerate(spreadsheets, start=1):
        node = spreadsheet.get("node", {})
        title = sanitize_filename(node_title(node))
        stem = f"{idx:03d}_{title}"
        xlsx_path = spreadsheet_base / f"{stem}.xlsx"
        raw_path = spreadsheet_base / f"{stem}.sheet.json"
        error_path = spreadsheet_base / f"{stem}.error.txt"
        spreadsheet_base.mkdir(parents=True, exist_ok=True)
        if spreadsheet.get("error"):
            error_path.write_text(f"{node_title(node)}\n\nERROR: {spreadsheet['error']}\n", encoding="utf-8")
            xlsx_value = None
        else:
            write_spreadsheet_xlsx(xlsx_path, spreadsheet)
            xlsx_value = str(xlsx_path)
        if include_raw:
            raw_path.write_text(json.dumps(spreadsheet, ensure_ascii=False, indent=2), encoding="utf-8")
        index["spreadsheets"].append(
            {
                "title": node_title(node),
                "path": node.get("_path"),
                "nodeId": node.get("nodeId"),
                "workbookId": spreadsheet.get("workbookId"),
                "sheets": len(spreadsheet.get("sheets", [])),
                "error": spreadsheet.get("error"),
                "xlsx": xlsx_value,
                "raw": str(raw_path) if include_raw else None,
            }
        )
    (base / "index.json").write_text(json.dumps(index, ensure_ascii=False, indent=2), encoding="utf-8")


def write_spreadsheet_xlsx(path: Path, spreadsheet: dict[str, Any]) -> None:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise DingTalkError("openpyxl is required to write .xlsx files. Install requirements.txt first.") from exc

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    used_titles: set[str] = set()
    sheet_entries = spreadsheet.get("sheets", [])
    if not isinstance(sheet_entries, list) or not sheet_entries:
        sheet = workbook.create_sheet("empty")
        sheet["A1"] = "No readable worksheets were returned by DingTalk."
    else:
        for entry in sheet_entries:
            if not isinstance(entry, dict):
                continue
            sheet_meta = entry.get("sheet") if isinstance(entry.get("sheet"), dict) else {}
            detail = entry.get("detail") if isinstance(entry.get("detail"), dict) else {}
            title = str(sheet_meta.get("name") or sheet_meta.get("title") or detail.get("name") or detail.get("title") or "Sheet").strip()
            worksheet = workbook.create_sheet(unique_sheet_title(title, used_titles))
            matrix = range_to_matrix(entry.get("range") if isinstance(entry.get("range"), dict) else {})
            if not matrix:
                worksheet["A1"] = ""
                continue
            for row_index, row in enumerate(matrix, start=1):
                for column_index, value in enumerate(row, start=1):
                    worksheet.cell(row=row_index, column=column_index, value=normalize_excel_value(value))
            auto_size_columns(worksheet)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def range_to_matrix(range_data: dict[str, Any]) -> list[list[Any]]:
    values = first_matrix(range_data, ("values",))
    formulas = first_matrix(range_data, ("formulas",))
    display_values = first_matrix(range_data, ("displayValues",))
    preferred = values or display_values or formulas or []
    if not preferred:
        return []
    row_count = max(
        len(preferred),
        len(formulas or []),
        len(display_values or []),
    )
    col_count = 0
    for matrix in (preferred, formulas, display_values):
        if matrix:
            col_count = max(col_count, max((len(row) for row in matrix if isinstance(row, list)), default=0))
    rows: list[list[Any]] = []
    for row_index in range(row_count):
        row: list[Any] = []
        for col_index in range(col_count):
            formula = matrix_value(formulas, row_index, col_index)
            value = matrix_value(values, row_index, col_index)
            display = matrix_value(display_values, row_index, col_index)
            if formula not in (None, ""):
                row.append(formula)
            elif value not in (None, ""):
                row.append(value)
            else:
                row.append(display)
        rows.append(row)
    return trim_empty_matrix(rows)


def first_matrix(value: Any, keys: tuple[str, ...]) -> list[list[Any]] | None:
    if isinstance(value, dict):
        for key in keys:
            item = value.get(key)
            if is_matrix(item):
                return item
        for item in value.values():
            found = first_matrix(item, keys)
            if found is not None:
                return found
    elif isinstance(value, list):
        for item in value:
            found = first_matrix(item, keys)
            if found is not None:
                return found
    return None


def is_matrix(value: Any) -> bool:
    return isinstance(value, list) and (not value or all(isinstance(row, list) for row in value))


def matrix_value(matrix: list[list[Any]] | None, row_index: int, col_index: int) -> Any:
    if not matrix or row_index >= len(matrix):
        return None
    row = matrix[row_index]
    if not isinstance(row, list) or col_index >= len(row):
        return None
    return row[col_index]


def trim_empty_matrix(matrix: list[list[Any]]) -> list[list[Any]]:
    while matrix and all(value in (None, "") for value in matrix[-1]):
        matrix.pop()
    if not matrix:
        return []
    max_col = 0
    for row in matrix:
        for idx, value in enumerate(row, start=1):
            if value not in (None, ""):
                max_col = max(max_col, idx)
    return [row[:max_col] for row in matrix]


def normalize_excel_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, dict):
        text = collect_text(value).strip()
        return text if text else json.dumps(value, ensure_ascii=False)
    if isinstance(value, list):
        text = collect_text(value).strip()
        return text if text else json.dumps(value, ensure_ascii=False)
    return str(value)


def unique_sheet_title(title: str, used_titles: set[str]) -> str:
    cleaned = re.sub(r"[\[\]:*?/\\]+", "_", title).strip() or "Sheet"
    cleaned = cleaned[:31]
    candidate = cleaned
    suffix = 2
    while candidate in used_titles:
        tail = f"_{suffix}"
        candidate = f"{cleaned[:31 - len(tail)]}{tail}"
        suffix += 1
    used_titles.add(candidate)
    return candidate


def auto_size_columns(worksheet: Any) -> None:
    for column_cells in worksheet.columns:
        letter = column_cells[0].column_letter
        max_length = 0
        for cell in column_cells[:200]:
            if cell.value is None:
                continue
            max_length = max(max_length, len(str(cell.value)))
        worksheet.column_dimensions[letter].width = min(max(max_length + 2, 8), 60)


def sanitize_filename(value: str) -> str:
    cleaned = re.sub(r'[<>:"/\\|?*\x00-\x1f]+', "_", value).strip(" .")
    return cleaned[:80] or "untitled"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Fetch DingTalk knowledge-base document nodes or blocks.")
    parser.add_argument("--access-token", help="Existing DingTalk OpenAPI accessToken. If omitted, appKey/appSecret generate a fresh token.")
    parser.add_argument("--app-key", default=os.getenv("DINGTALK_APP_KEY"), help="DingTalk appKey.")
    parser.add_argument("--app-secret", default=os.getenv("DINGTALK_APP_SECRET"), help="DingTalk appSecret.")
    parser.add_argument("--corp-id", default=os.getenv("DINGTALK_CORP_ID"), help="DingTalk corpId, kept for diagnostics.")
    parser.add_argument("--operator-union-id", default=os.getenv("DINGTALK_OPERATOR_UNION_ID"), help="Operator unionId.")
    parser.add_argument("--lookup-userid-by-mobile", help="Resolve userid from mobile number, then exit.")
    parser.add_argument("--lookup-unionid-by-userid", help="Resolve unionId from DingTalk userid, then exit.")
    parser.add_argument("--url", help="DingTalk alidocs URL to resolve and read.")
    parser.add_argument("--workspace-id", help="Known DingTalk knowledge-base workspaceId.")
    parser.add_argument("--root-node-id", help="Known DingTalk knowledge-base root nodeId.")
    parser.add_argument("--doc-key", help="Known document docKey or nodeId to read blocks.")
    parser.add_argument("--node-id", help="Known folder/root nodeId to list, or document nodeId to use as docKey.")
    parser.add_argument("--crawl", action="store_true", help="Recursively fetch document blocks under a workspace/root/folder.")
    parser.add_argument("--list-root", action="store_true", help="List nodes under the mineWorkspaces rootNodeId.")
    parser.add_argument("--list-children", action="store_true", help="List children under --node-id.")
    parser.add_argument("--format", choices=("markdown", "json"), default="markdown", help="Output format for document blocks.")
    parser.add_argument("--output", help="Write converted output to this file.")
    parser.add_argument("--output-dir", help="Directory for recursive crawl outputs.")
    parser.add_argument("--raw-output", help="Write raw JSON response to this file.")
    parser.add_argument("--include-raw", action="store_true", help="Write raw JSON beside each crawled Markdown file.")
    parser.add_argument("--max-results", type=int, default=50, help="Page size for list/read APIs.")
    parser.add_argument("--depth-limit", type=int, default=8, help="Maximum folder recursion depth for --crawl.")
    parser.add_argument("--limit-docs", type=int, help="Maximum number of documents to read during --crawl.")
    parser.add_argument("--include-spreadsheets", "--include-tables", action="store_true", help="Also export online spreadsheets (.axls) to .xlsx during --crawl.")
    parser.add_argument("--limit-spreadsheets", type=int, help="Maximum number of online spreadsheets to export during --crawl.")
    return parser.parse_args()


def main() -> int:
    repo_root = Path(__file__).resolve().parents[3]
    load_dotenv(repo_root)
    args = parse_args()

    token = resolve_access_token(args)

    if args.lookup_userid_by_mobile:
        result = get_userid_by_mobile(token, args.lookup_userid_by_mobile)
        write_output(args.output, json.dumps(result, ensure_ascii=False, indent=2))
        return 0

    if args.lookup_unionid_by_userid:
        result = get_unionid_by_userid(token, args.lookup_unionid_by_userid)
        write_output(args.output, json.dumps(result, ensure_ascii=False, indent=2))
        return 0

    operator_id = args.operator_union_id or os.getenv("DINGTALK_OPERATOR_UNION_ID")
    if not operator_id:
        raise DingTalkError("Missing operator unionId. Use --lookup-userid-by-mobile or --lookup-unionid-by-userid to resolve it first.")
    workspace_id = args.workspace_id
    root_node_id_from_url: str | None = None
    space_url_node: dict[str, Any] | None = None
    space_url_diagnostics: list[str] = []
    if args.url and extract_workspace_id_from_url(args.url):
        workspace_id_from_url, root_node_id_from_url, space_url_node, space_url_diagnostics = resolve_space_url(token, operator_id, args.url)
        workspace_id = workspace_id or workspace_id_from_url

    if args.list_root:
        workspaces = get_mine_workspaces(token, operator_id)
        root_node_id = None
        if isinstance(workspaces, dict):
            workspace = workspaces.get("workspace") if isinstance(workspaces.get("workspace"), dict) else workspaces
            root_node_id = workspace.get("rootNodeId") if isinstance(workspace, dict) else None
        if not root_node_id:
            raise DingTalkError(f"rootNodeId missing in mineWorkspaces response: {json.dumps(workspaces, ensure_ascii=False)[:500]}")
        nodes = list_nodes(token, operator_id, root_node_id, args.max_results)
        raw = {"mineWorkspaces": workspaces, "nodes": nodes}
        text = json.dumps(raw, ensure_ascii=False, indent=2)
        write_output(args.output, text)
        return 0

    if workspace_id and not args.crawl and not args.list_children and not args.doc_key and not args.node_id:
        if root_node_id_from_url:
            nodes = list_nodes(token, operator_id, root_node_id_from_url, args.max_results)
        else:
            try:
                nodes = list_workspace_nodes(token, operator_id, workspace_id, args.max_results)
            except DingTalkError as exc:
                if space_url_diagnostics:
                    raise DingTalkError(f"{exc} Space URL diagnostics: {' | '.join(space_url_diagnostics)}") from exc
                raise
        raw = {"workspaceId": workspace_id, "rootNodeId": root_node_id_from_url, "spaceUrlNode": space_url_node, "nodes": nodes}
        write_output(args.output, json.dumps(raw, ensure_ascii=False, indent=2))
        return 0

    if args.list_children:
        if not args.node_id:
            raise DingTalkError("--list-children requires --node-id.")
        nodes = list_nodes(token, operator_id, args.node_id, args.max_results)
        text = json.dumps({"nodes": nodes}, ensure_ascii=False, indent=2)
        write_output(args.output, text)
        return 0

    if args.crawl:
        root_nodes: list[dict[str, Any]]
        if args.root_node_id:
            root_nodes = list_nodes(token, operator_id, args.root_node_id, args.max_results)
        elif root_node_id_from_url:
            root_nodes = list_nodes(token, operator_id, root_node_id_from_url, args.max_results)
        elif args.node_id:
            root_nodes = [{"nodeId": args.node_id, "name": args.node_id, "type": "folder"}]
        elif workspace_id:
            try:
                root_nodes = list_workspace_nodes(token, operator_id, workspace_id, args.max_results)
            except DingTalkError as exc:
                if space_url_diagnostics:
                    raise DingTalkError(f"{exc} Space URL diagnostics: {' | '.join(space_url_diagnostics)}") from exc
                raise
        elif args.url and not extract_workspace_id_from_url(args.url):
            root_nodes = [query_node_by_url(token, operator_id, args.url)]
        else:
            raise DingTalkError("--crawl requires --workspace-id, a space --url, --root-node-id, --node-id, or document --url.")

        crawled_nodes = crawl_nodes(token, operator_id, root_nodes, args.max_results, args.depth_limit)
        documents = fetch_documents(token, operator_id, crawled_nodes, args.max_results, args.limit_docs)
        spreadsheets = (
            fetch_spreadsheets(token, operator_id, crawled_nodes, args.limit_spreadsheets)
            if args.include_spreadsheets
            else []
        )
        aitable_nodes = [node for node in crawled_nodes if is_aitable_node(node)]
        summary = {
            "workspaceId": workspace_id,
            "totalNodes": len(crawled_nodes),
            "documents": len(documents),
            "failedDocuments": sum(1 for doc in documents if doc.get("error")),
            "spreadsheets": len(spreadsheets),
            "failedSpreadsheets": sum(1 for item in spreadsheets if item.get("error")),
            "aitableNodes": len(aitable_nodes),
        }
        if args.output_dir:
            write_crawl_outputs(args.output_dir, documents, spreadsheets, args.include_raw)
            summary["outputDir"] = args.output_dir
        if args.raw_output:
            write_output(
                args.raw_output,
                json.dumps(
                    {
                        "summary": summary,
                        "nodes": crawled_nodes,
                        "documents": documents,
                        "spreadsheets": spreadsheets,
                        "aitableNodes": aitable_nodes,
                    },
                    ensure_ascii=False,
                    indent=2,
                ),
            )
        write_output(args.output, json.dumps(summary, ensure_ascii=False, indent=2))
        return 0

    node: dict[str, Any] = {}
    if args.url and not extract_workspace_id_from_url(args.url):
        node = query_node_by_url(token, operator_id, args.url)
    doc_key = args.doc_key or args.node_id or extract_doc_key(node, args.url)
    if not doc_key:
        raise DingTalkError("No document key found. Provide --doc-key, --node-id, or a resolvable --url.")

    raw = read_blocks(token, operator_id, doc_key, args.max_results)
    if args.raw_output:
        write_output(args.raw_output, json.dumps({"node": node, **raw}, ensure_ascii=False, indent=2))

    if args.format == "json":
        write_output(args.output, json.dumps({"node": node, **raw}, ensure_ascii=False, indent=2))
    else:
        markdown = blocks_to_markdown(raw["blocks"])
        write_output(args.output, markdown)
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except DingTalkError as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        raise SystemExit(1)
