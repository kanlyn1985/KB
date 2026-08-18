#!/usr/bin/env python3
"""LLM 落位结果 → Excel 导出（skeleton_v0.3_landing.xlsx）。

数据源：llm_landing/reland/reland_records.jsonl（225420 条完整落位）
        llm_landing/reland/reland_review.jsonl（48483 条复核队列）
骨架：  skeleton_v0.2.json（177 节点）

Sheet1「节点落位统计」：177 节点 × 归属单元数/复核数（发现过载/空节点）
Sheet2「落位明细」：全部落位记录（文档/单元/内容/节点/置信度/理由）
Sheet3「复核队列」：待人工处理项
"""

from __future__ import annotations

import json
import sys
from collections import Counter
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
TREE = ROOT / "docs" / "ontology" / "tree_skeleton"
SKELETON = TREE / "skeleton_v0.2.json"
RECORDS = TREE / "llm_landing" / "reland" / "reland_records.jsonl"
REVIEW = TREE / "llm_landing" / "reland" / "reland_review.jsonl"
OUT_XLSX = TREE / "skeleton_v0.3_landing.xlsx"

LAYER_COLORS = {
    "P": "E2EFDA", "F": "DDEBF7", "L": "FFF2CC",
    "R": "FCE4D6", "G": "EDEDED", "M": "D9E1F2", "Q": "F8CBAD",
}


def main() -> int:
    sys.path.insert(0, str(ROOT / ".venv-paddle" / "lib" / "python3.12" / "site-packages"))
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    data = json.loads(SKELETON.read_text(encoding="utf-8"))
    nodes = data["nodes"]
    node_map = {n["id"]: n for n in nodes}

    # 统计落位
    node_cnt = Counter()
    review_cnt = Counter()
    rows_all = []
    with RECORDS.open(encoding="utf-8") as f:
        for line in f:
            if not line.strip():
                continue
            r = json.loads(line)
            rows_all.append(r)
            if r.get("node_id"):
                node_cnt[r["node_id"]] += 1
            else:
                review_cnt[r.get("doc", "")] += 1

    review_rows = []
    if REVIEW.exists():
        with REVIEW.open(encoding="utf-8") as f:
            for line in f:
                if line.strip():
                    review_rows.append(json.loads(line))

    wb = openpyxl.Workbook()
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin = Border(*[Side(style="thin", color="D9D9D9")] * 4)

    def style_header(ws, ncols):
        for col in range(1, ncols + 1):
            c = ws.cell(row=1, column=col)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal="center", vertical="center")

    # === Sheet1: 节点落位统计 ===
    ws1 = wb.active
    ws1.title = "节点落位统计"
    headers1 = ["节点ID", "层", "类型", "节点名称", "父节点ID", "归属单元数", "占总数比", "备注"]
    ws1.append(headers1)
    style_header(ws1, len(headers1))
    total_units = len(rows_all)
    total_assigned = sum(node_cnt.values())
    for n in nodes:
        c = node_cnt.get(n["id"], 0)
        pct = f"{c / total_units * 100:.2f}%" if total_units else "-"
        note = ""
        if c == 0:
            note = "空节点（无内容落位）"
        elif c > 5000:
            note = "过载（>5000，建议细分）"
        row = [n["id"], n["layer"], n["type"], n["name"], n.get("parent", ""), c, pct, note]
        ws1.append(row)
        fill = PatternFill("solid", fgColor=LAYER_COLORS.get(n["layer"], "FFFFFF"))
        for col in range(1, len(headers1) + 1):
            cell = ws1.cell(row=ws1.max_row, column=col)
            cell.fill = fill
            cell.border = thin
            cell.alignment = Alignment(vertical="top", wrap_text=(col in (4, 8)))
    ws1.append([])
    ws1.append(["总计", "", "", "", "", total_assigned,
                f"{total_assigned / total_units * 100:.1f}%" if total_units else "-",
                f"总单元 {total_units} / 归属 {total_assigned} / 未归属 {total_units - total_assigned}"])
    for i, w in enumerate([22, 6, 12, 45, 14, 12, 10, 20], 1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    ws1.freeze_panes = "A2"

    # === Sheet2: 落位明细 ===
    ws2 = wb.create_sheet("落位明细")
    headers2 = ["文档", "单元ID", "类型", "内容", "节点ID", "节点名称", "置信度", "LLM理由", "脚本规则"]
    ws2.append(headers2)
    style_header(ws2, len(headers2))
    # 明细用 write_only 模式太慢；直接逐行 append（22 万行，openpyxl 约几分钟）
    for r in rows_all:
        ws2.append([
            r.get("doc", ""), r.get("unit_id", ""), r.get("unit_type", ""),
            r.get("text", ""), r.get("node_id") or "", r.get("node_name") or "",
            r.get("conf"), r.get("llm_reason", ""), r.get("rule", ""),
        ])
    for i, w in enumerate([40, 26, 8, 50, 18, 40, 8, 40, 14], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"

    # === Sheet3: 复核队列 ===
    ws3 = wb.create_sheet("复核队列")
    headers3 = ["文档", "单元ID", "类型", "内容", "原因"]
    ws3.append(headers3)
    style_header(ws3, len(headers3))
    for r in review_rows:
        ws3.append([r.get("doc", ""), r.get("unit_id", ""), r.get("unit_type", ""),
                    r.get("text", ""), r.get("reason", "")])
    for i, w in enumerate([40, 26, 8, 50, 40], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    ws3.freeze_panes = "A2"

    wb.save(OUT_XLSX)
    print(f"✅ Excel 已导出: {OUT_XLSX}")
    print(f"   Sheet1「节点落位统计」: {len(nodes)} 节点 × 归属数")
    print(f"   Sheet2「落位明细」: {len(rows_all)} 行")
    print(f"   Sheet3「复核队列」: {len(review_rows)} 行")
    empty = [n["id"] for n in nodes if node_cnt.get(n["id"], 0) == 0]
    print(f"   空节点: {len(empty)} 个 | 过载节点: {sum(1 for n in nodes if node_cnt.get(n['id'], 0) > 5000)} 个")
    return 0


if __name__ == "__main__":
    sys.exit(main())
