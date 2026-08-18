"""内容单元提取器：把文档解析成可落位的内容单元。

内容单元 = 文档里可独立落位的最小信息块：
  - md: 按标题层级切分段落块 + 提取表格行
  - docx: 段落 + 表格（zipfile + XML 解析，无 python-docx 依赖）
  - xlsx: 行（openpyxl，若可用）
  - pdf: 文本层（pdftotext 外部命令）

输出单元结构：
  {doc, unit_type, section, text, table_ref, line_no, unit_id}

unit_type 初判（内容模式 → 单元类型）：
  table_row / strategy_para / requirement / clause / component /
  process / experience / para
"""

from __future__ import annotations

import json
import re
import subprocess
import sys
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

W = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
ROOT = Path(__file__).resolve().parents[2]
SKELETON = ROOT / "docs" / "ontology" / "tree_skeleton" / "skeleton_v0.2.json"
MANIFEST = ROOT / "docs" / "ontology" / "tree_skeleton" / "doc_manifest.json"

# 表格行特征：含 | 分隔（md 表格）或多列（xlsx）
TABLE_LINE = re.compile(r"^\s*\|.+\|\s*$")

# ── 噪声白名单（v0.4.0：让归属数字诚实）────────────────────────────
# 这些文本不是知识单元，不计归属也不进复核队列：
#   SN 序列号 / 附件引用 / 图片占位 / HTML 注释占位 / 纯路径 / 版本号 / 无意义标题
NOISE_SN = re.compile(
    r"^(SN[:：]?\s*[A-Za-z0-9\-]{6,}|[A-Za-z0-9]{2,}-\d{4,}[A-Za-z0-9\-]{6,}$"
    r"|V\d(\.\d)*[-_][A-Z]{2,3}[-_]\d{6,})"
)
NOISE_PATH = re.compile(
    r"^([A-Za-z]:[\\/]|N:\\\\|\\\\|/mnt/|/home/|http[s]?://|ftp://|\\\\192\.)"
)
NOISE_ATTACH = re.compile(r"^(\[附件:|!\[|<!-- unsupported DingTalk block|dingtalk-resource://)")
NOISE_TABLE_HEADER = re.compile(
    r"^(\|?\s*(ID|序号|No\.?|Item|Date|Author|版本|时间|备注|说明|V\d\.\d)[\s|]*(\|?\s*)*$|^V\d\.\d\s*\|\s*(Date|Author|EVT|初稿))",
    re.IGNORECASE,
)
NOISE_TINY = re.compile(
    r"^(结论[:：]?|建议[:：]?|总结[:：]?|概述[:：]?|背景[:：]?|前言[:：]?|附录|附表|"
    r"样例参考[:：]?|参考[:：]?|备注[:：]?|注意[:：]?|说明[:：]?|方案[:：]?|"
    r"验证方案[:：]?|原因[:：]?|改善方案[:：]?|Figure\s+\d+(\.\d+)*|"
    r"第[一二三四五六七八九十\d]+[章节部分]|附录[A-Z]?|Appendix|"
    r"【[^】]{1,12}】)$"
)


def is_noise_text(text: str) -> bool:
    """判定文本是否为噪声（非知识单元）。"""
    t = text.strip()
    if not t:
        return True
    if NOISE_ATTACH.search(t):
        return True
    if NOISE_SN.match(t):
        return True
    if NOISE_PATH.match(t):
        return True
    if NOISE_TABLE_HEADER.match(t):
        return True
    if len(t) <= 2:
        return True
    # 纯数字/纯标点/纯序列
    if re.fullmatch(r"[\d\s\-–—./%()【】\[\]·,，。;；:：|]*", t):
        return True
    if NOISE_TINY.fullmatch(t):
        return True
    return False


def extract_md(path: Path) -> list[dict]:
    path = Path(path)
    units = []
    text = path.read_text(encoding="utf-8", errors="replace")
    lines = text.splitlines()
    section = "顶层"
    section_stack: list[str] = []
    para_buf: list[str] = []
    line_no = 0
    table_buf: list[str] = []

    def flush_para():
        nonlocal para_buf
        if para_buf:
            t = " ".join(x.strip() for x in para_buf if x.strip())
            if t:
                units.append({"unit_type": "para", "section": section,
                              "text": t[:500], "line_no": line_no})
            para_buf = []

    def flush_table():
        nonlocal table_buf
        if table_buf:
            # 表头 + 数据行
            header = table_buf[0] if table_buf else ""
            for i, row in enumerate(table_buf[1:], 1):
                cells = [c.strip() for c in row.strip("|").split("|")]
                if len(cells) >= 2 and cells[0] and cells[0] != "---":
                    units.append({
                        "unit_type": "table_row",
                        "section": section,
                        "text": " | ".join(cells),
                        "table_ref": f"{path.name}:表{header[:40]}行{i}",
                        "line_no": line_no,
                    })
            table_buf = []

    for raw in lines:
        line_no += 1
        line = raw.strip()
        if not line:
            flush_para()
            continue
        if TABLE_LINE.match(line):
            flush_para()
            table_buf.append(line)
            continue
        if table_buf:
            flush_table()
            # 噪声：图片占位/OCR推理占位/模板占位（决策4：收集待人工复核）
            if line.startswith("![") or "推理补充" in line or "需人工核对" in line \
               or line.startswith("（描述") or line.startswith("（若") or line.startswith("（待") \
               or "待补充" in line or "待补充:" in line \
               or line.startswith("> **转换日期") or line.startswith("> **上次更新") \
               or line.startswith("> **创建人") or line.startswith("> **来源") \
               or line.startswith("> 来源:") or line.startswith("> 日期") \
               or line.strip() in ("---", "```", "---", ">", "```yaml", "```json"):
                units.append({"unit_type": "noise", "section": section,
                              "text": line[:200], "line_no": line_no})
                continue
        m = re.match(r"^(#{1,6})\s+(.+)$", line)
        if m:
            flush_para()
            lvl = len(m.group(1))
            title = m.group(2).strip()
            while len(section_stack) >= lvl:
                section_stack.pop()
            section_stack.append(title)
            section = "/".join(section_stack)
            # 标题本身也是内容单元（策略/组件/经验文档的标题是节点名线索）
            units.append({"unit_type": "heading", "section": section,
                          "text": title, "line_no": line_no})
            continue
        para_buf.append(line)
    flush_para()
    flush_table()

    # 单元类型细化（内容模式 → 类型）
    for u in units:
        u["unit_type"] = classify(u["text"], u["unit_type"], path)
        if is_noise_text(u["text"]):
            u["unit_type"] = "noise"
        u["doc"] = str(path)
        u["unit_id"] = f"{path.stem[:20]}#{u['line_no']}"
    return units


def classify(text: str, fallback: str, path: Path) -> str:
    """内容模式 → 单元类型（确定性规则初判，LLM 候选在后置环节）"""
    t = text.strip()
    # 表格行：按列内容判定
    if fallback == "table_row":
        return "table_row"
    # 需求样式（编号+要求/应/须）
    if re.match(r"^(R-|SWRD|EVT-R|MI-MD|SW4-|FUNC|REQ)", t) or "应满足" in t or "需满足" in t or "必须" in t:
        return "requirement"
    # 标准条款样式（条款号）
    if re.match(r"^\d+(\.\d+)*\s", t) and ("要求" in t or "试验" in t or "条款" in t):
        return "clause"
    # 经验/问题样式
    if any(k in t for k in ("踩坑", "问题", "故障", "根因", "原因分析", "教训", "复盘", "FAQ")):
        return "experience"
    # 流程样式
    if any(k in t for k in ("流程", "步骤", "阶段", "SOP", "工艺")):
        return "process"
    # 策略样式
    if "策略" in t or "算法" in t or "逻辑" in t:
        return "strategy"
    # 组件样式
    if any(k in t for k in ("组件", "模块", "SW-C", "详细设计")):
        return "component"
    return fallback


# docx 元数据关键词（封面/文件信息区）
DOCX_META_KEYS = ("文件标识", "当前版本", "文件状态", "批准", "审核", "作者",
                  "完成日期", "版权", "变更记录", "变更原因", "影响章节",
                  "变更内容", "目录", "引文", "表1-1", "No.", "Document", "Version", "Note")
# docx 信号定义字段
SIGNAL_KEYS = ("信号名称", "信号描述", "信号单位", "单位", "精度", "信号范围",
               "初始值", "上报逻辑", "前提条件", "触发条件", "执行动作")


def extract_docx(path: Path) -> list[dict]:
    path = Path(path)
    units = []
    try:
        with zipfile.ZipFile(path) as z:
            xml = z.read("word/document.xml")
        root = ET.fromstring(xml)
        body = root.find(W + "body")
        if body is None:
            body = root
        idx = 0
        for child in body:
            tag = child.tag
            if tag == W + "p":
                idx += 1
                texts = [t.text or "" for t in child.iter(W + "t")]
                line = "".join(texts).strip()
                # 过滤空段（含零宽空格）
                if not line or line.strip("\u200b\u200c\u200d\ufeff ").strip() == "":
                    continue
                utype = "para"
                if any(k in line for k in DOCX_META_KEYS) and len(line) < 60:
                    utype = "meta"
                elif re.search(r"\bSW4-\d+", line):
                    utype = "requirement"
                elif any(k in line for k in SIGNAL_KEYS):
                    utype = "signal"
                elif re.match(r"^#{1,3}\s", line) or (line and len(line) < 40 and not line.endswith(("。", ".", "，"))):
                    utype = "heading"
                units.append({"doc": str(path), "unit_id": f"{path.stem[:20]}#p{idx}",
                              "unit_type": utype, "section": "",
                              "text": line[:500], "line_no": idx})
            elif tag == W + "tbl":
                for r_i, tr in enumerate(child.iter(W + "tr")):
                    idx += 1
                    cells = []
                    for tc in tr.iter(W + "tc"):
                        t = "".join(x.text or "" for x in tc.iter(W + "t")).strip()
                        if t:
                            cells.append(t)
                    if cells:
                        units.append({"doc": str(path), "unit_id": f"{path.stem[:20]}#tbl{idx}",
                                      "unit_type": "table_row", "section": "",
                                      "text": " | ".join(cells)[:500], "line_no": idx})
    except Exception as e:  # noqa: BLE001
        print(f"docx 解析失败 {path.name}: {e}", file=sys.stderr)
    return units


def extract_xlsx(path: Path) -> list[dict]:
    path = Path(path)
    units = []
    try:
        sys.path.insert(0, str(ROOT / ".venv-paddle" / "lib" / "python3.12" / "site-packages"))
        import openpyxl  # type: ignore
        wb = openpyxl.load_workbook(path, data_only=True)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            rows = list(ws.iter_rows(values_only=True))
            header = [str(v).strip() if v else "" for v in rows[0]]
            # 需求表检测：列头含 ID 且第二列是"类型"
            is_req_table = len(header) >= 2 and "ID" in header and "类型" in header
            type_col = header.index("类型") if "类型" in header else None
            for i, row in enumerate(rows[1:], 2):
                vals = [str(v).strip() for v in row if v is not None and str(v).strip()]
                if not vals:
                    continue
                text = " | ".join(vals)[:500]
                if is_req_table and type_col is not None and type_col < len(row):
                    rtype = str(row[type_col]).strip()
                    # 按类型列映射单元类型
                    if rtype == "标题":
                        utype = "heading"
                    elif rtype == "系统需求" or "需求" in rtype:
                        utype = "requirement"
                    elif rtype == "CAN矩阵":
                        utype = "signal"
                    elif rtype == "故障列表":
                        utype = "fault"
                    else:
                        utype = "table_row"
                else:
                    utype = "table_row"
                # 噪声白名单：SN/编号/占位/纯数字行 → 不计入知识单元
                if is_noise_text(text):
                    continue
                units.append({"doc": str(path), "unit_id": f"{path.stem[:20]}#{sheet}#{i}",
                              "unit_type": utype, "section": sheet,
                              "text": text, "line_no": i})
    except Exception as e:  # noqa: BLE001
        print(f"xlsx 解析失败 {path.name}: {e}", file=sys.stderr)
    return units


def extract_pdf(path: Path) -> list[dict]:
    path = Path(path)
    units = []
    try:
        r = subprocess.run(["pdftotext", "-enc", "UTF-8", str(path), "-"],
                           capture_output=True, text=True, timeout=120)
        text = r.stdout
        lines = [l for l in text.splitlines() if l.strip()]
        # 按条款号切分（标准类 PDF）
        buf: list[str] = []
        cur_section = "顶层"
        for i, line in enumerate(lines):
            m = re.match(r"^(\d+(?:\.\d+)*)\s*(.+)$", line.strip())
            if m and (len(m.group(1)) <= 3):
                if buf:
                    units.append({"doc": str(path), "unit_id": f"{path.stem[:20]}#{i}",
                                  "unit_type": classify(" ".join(buf), "clause", path),
                                  "section": cur_section, "text": " ".join(buf)[:500], "line_no": i})
                cur_section = f"{m.group(1)} {m.group(2)[:40]}"
                buf = [line.strip()]
            else:
                buf.append(line.strip())
        if buf:
            units.append({"doc": str(path), "unit_id": f"{path.stem[:20]}#end",
                          "unit_type": "clause", "section": cur_section,
                          "text": " ".join(buf)[:500], "line_no": len(lines)})
    except Exception as e:  # noqa: BLE001
        print(f"pdf 解析失败 {path.name}: {e}", file=sys.stderr)
    return units


def main() -> None:
    import argparse
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--category", default=None, help="只处理某 Athena 分类（如 30_产品平台知识）")
    parser.add_argument("--limit", type=int, default=0, help="限制文档数（0=全部）")
    parser.add_argument("--output", default=str(ROOT / "docs" / "ontology" / "tree_skeleton" / "units_sample.json"))
    args = parser.parse_args()

    manifest = json.loads(MANIFEST.read_text(encoding="utf-8"))
    docs = [d for d in manifest["docs"] if d["ext"] in {".md", ".docx", ".xlsx", ".pdf"}]
    if args.category:
        docs = [d for d in docs if d.get("category") == args.category]
    if args.limit:
        docs = docs[: args.limit]

    all_units = []
    per_doc = {}
    for d in docs:
        p = Path(d["path"])
        if p.suffix == ".md":
            units = extract_md(p)
        elif p.suffix == ".docx":
            units = extract_docx(p)
        elif p.suffix == ".xlsx":
            units = extract_xlsx(p)
        elif p.suffix == ".pdf":
            units = extract_pdf(p)
        else:
            units = []
        all_units.extend(units)
        per_doc[p.name] = len(units)

    out = Path(args.output)
    out.write_text(json.dumps({
        "docs_processed": len(docs),
        "total_units": len(all_units),
        "per_doc": per_doc,
        "units": all_units[:2000],
    }, ensure_ascii=False, indent=1), encoding="utf-8")

    from collections import Counter
    print(f"处理文档 {len(docs)} 份，提取内容单元 {len(all_units)} 个")
    print("单元类型分布:", dict(Counter(u["unit_type"] for u in all_units)))


if __name__ == "__main__":
    main()
