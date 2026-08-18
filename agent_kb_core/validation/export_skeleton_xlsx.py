"""骨架 v0.2 → Excel 导出：完整信息，可随时查看。

输出 docs/ontology/tree_skeleton/skeleton_v0.2.xlsx
Sheet1「骨架节点」：每节点一行（ID/层/类型/名称/父节点/层级路径/层级深度/子节点数）
Sheet2「分支概览」：按层统计
Sheet3「树形视图」：缩进展示树结构（人工浏览用）
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
SKELETON = ROOT / "docs" / "ontology" / "tree_skeleton" / "skeleton_v0.2.json"
OUT_XLSX = ROOT / "docs" / "ontology" / "tree_skeleton" / "skeleton_v0.2.xlsx"


def build_tree(nodes: list[dict]) -> dict[str, dict]:
    """构建节点树：id → node（含 children/depth/path）"""
    tree = {n["id"]: {**n, "children": [], "depth": 0, "path": "", "child_count": 0} for n in nodes}
    roots = []
    for n in nodes:
        pid = n.get("parent")
        if pid and pid in tree:
            tree[pid]["children"].append(n["id"])
        else:
            roots.append(n["id"])
    # 递归算 depth/path/child_count
    def walk(nid: str, depth: int, path: str):
        node = tree[nid]
        node["depth"] = depth
        node["path"] = path + " / " + node["name"] if path else node["name"]
        node["child_count"] = len(node["children"])
        for cid in node["children"]:
            walk(cid, depth + 1, node["path"])
    for rid in roots:
        walk(rid, 0, "")
    return tree, roots


def main() -> None:
    import argparse
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--no-sync-xlsx", action="store_true", help="兼容参数（无实际作用，防递归）")
    args = parser.parse_args()

    sys.path.insert(0, str(ROOT / ".venv-paddle" / "lib" / "python3.12" / "site-packages"))
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    data = json.loads(SKELETON.read_text(encoding="utf-8"))
    nodes = data["nodes"]
    tree, roots = build_tree(nodes)

    wb = openpyxl.Workbook()

    # === Sheet1: 骨架节点（每节点一行）===
    ws1 = wb.active
    ws1.title = "骨架节点"
    headers = ["节点ID", "层", "类型", "节点名称", "父节点ID", "父节点名称",
               "层级深度", "层级路径", "子节点数", "tree_version"]
    ws1.append(headers)
    # 表头样式
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin = Border(*[Side(style="thin", color="D9D9D9")] * 4)
    for col in range(1, len(headers) + 1):
        cell = ws1.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 层颜色
    layer_colors = {
        "P": "E2EFDA",  # 绿
        "F": "DDEBF7",  # 蓝
        "L": "FFF2CC",  # 黄
        "R": "FCE4D6",  # 橙
        "G": "EDEDED",  # 灰
        "M": "D9E1F2",  # 浅蓝
        "Q": "F8CBAD",  # 浅红
    }
    for n in nodes:
        t = tree[n["id"]]
        row = [n["id"], n["layer"], n["type"], n["name"],
               n.get("parent", ""), tree[n.get("parent", "")]["name"] if n.get("parent") in tree else "",
               t["depth"], t["path"], t["child_count"], data["tree_version"]]
        ws1.append(row)
        fill = PatternFill("solid", fgColor=layer_colors.get(n["layer"], "FFFFFF"))
        for col in range(1, len(headers) + 1):
            ws1.cell(row=ws1.max_row, column=col).fill = fill
            ws1.cell(row=ws1.max_row, column=col).border = thin
            ws1.cell(row=ws1.max_row, column=col).alignment = Alignment(vertical="top", wrap_text=True)

    # 列宽
    widths = [22, 6, 12, 45, 14, 30, 8, 60, 8, 12]
    for i, w in enumerate(widths, 1):
        ws1.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws1.freeze_panes = "A2"

    # === Sheet2: 分支概览 ===
    ws2 = wb.create_sheet("分支概览")
    ws2.append(["层", "节点数", "根节点", "说明"])
    for col in range(1, 5):
        c = ws2.cell(row=1, column=col)
        c.fill = header_fill; c.font = header_font; c.alignment = Alignment(horizontal="center")
    layer_desc = {
        "P": "物理分解树（主骨架）：硬件/软件/标定/接口/知识",
        "F": "功能分解树（黑盒能力）：OBC/DCDC/系统功能",
        "L": "逻辑组件树（白盒组件）：策略/控制/状态/故障/通信",
        "R": "需求追踪层（贯穿各层）：性能/保护/安全/EMC/环境/可靠性/软硬件/接口/标准",
        "G": "过程维度（正交）：开发过程/生产过程/ASPICE/开发方法",
        "M": "实例层（型号 vs 实例）：CCU型号/曼岛/VAVE/G5/SW4.0",
        "Q": "质量/经验域：问题记录/经验教训/失效模式",
    }
    from collections import Counter
    layer_counts = Counter(n["layer"] for n in nodes)
    for layer in ["P", "F", "L", "R", "G", "M", "Q"]:
        root_names = [tree[rid]["name"] for rid in roots if tree[rid]["layer"] == layer]
        ws2.append([layer, layer_counts.get(layer, 0), " / ".join(root_names), layer_desc.get(layer, "")])
        fill = PatternFill("solid", fgColor=layer_colors.get(layer, "FFFFFF"))
        for col in range(1, 5):
            ws2.cell(row=ws2.max_row, column=col).fill = fill
            ws2.cell(row=ws2.max_row, column=col).border = thin
    ws2.append([])
    ws2.append(["总计", len(nodes), "", f"tree_version={data['tree_version']}  branches={data['branches']}"])
    for i, w in enumerate([6, 8, 40, 60], 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"

    # === Sheet3: 树形视图（缩进展示）===
    ws3 = wb.create_sheet("树形视图")
    ws3.append(["缩进展示", "节点ID", "层", "类型", "名称"])
    for col in range(1, 6):
        c = ws3.cell(row=1, column=col)
        c.fill = header_fill; c.font = header_font; c.alignment = Alignment(horizontal="center")

    def render_tree(nid: str, indent: str = ""):
        t = tree[nid]
        prefix = indent + ("├─ " if indent else "")
        ws3.append([prefix + t["name"], nid, t["layer"], t["type"], t["name"]])
        fill = PatternFill("solid", fgColor=layer_colors.get(t["layer"], "FFFFFF"))
        for col in range(1, 6):
            ws3.cell(row=ws3.max_row, column=col).fill = fill
            ws3.cell(row=ws3.max_row, column=col).border = thin
        for i, cid in enumerate(t["children"]):
            is_last = i == len(t["children"]) - 1
            child_indent = indent + ("│  " if indent and not is_last else "   " if indent else "")
            render_tree(cid, child_indent)

    for rid in roots:
        render_tree(rid)
    ws3.column_dimensions["A"].width = 70
    for i, w in enumerate([22, 6, 12], 2):
        ws3.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws3.freeze_panes = "A2"

    wb.save(OUT_XLSX)
    print(f"✅ Excel 已导出: {OUT_XLSX}")
    print(f"   Sheet1「骨架节点」: {len(nodes)} 行（每节点一行，10 列完整信息）")
    print(f"   Sheet2「分支概览」: 7 层统计")
    print(f"   Sheet3「树形视图」: 缩进树结构")
    print(f"   节点数: {len(nodes)}  层: P{layer_counts['P']} F{layer_counts['F']} L{layer_counts['L']} "
          f"R{layer_counts['R']} G{layer_counts['G']} M{layer_counts['M']} Q{layer_counts['Q']}")


if __name__ == "__main__":
    main()
